from django.shortcuts import redirect, render, get_object_or_404
from django.views import View
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Formula, FunksionalData
import types
import math
from rest_framework import status
from django.utils.text import slugify
from django.http import HttpResponse
from openpyxl import load_workbook
import os 
from django.conf import settings
from datetime import datetime


class FormulaPage(View):
    def get(self, request):
        context = {
            'formulas': Formula.objects.all()
        }
        return render(request, 'formula-list.html', context)
    

class FormulaDetail(View):
    def get(self, request, pk):
        formula = get_object_or_404(Formula, pk=pk)
        other_formulas = Formula.objects.exclude(pk=pk)
        
        context = {
            'formula': formula,
            'formulas': other_formulas
        }
        return render(request, 'formula_detail.html', context)
    
    def post(self, request, pk):
        args = request.POST

        formula = Formula.objects.get(pk=pk)
        code = formula.code
        
        data = dict()
        for i in formula.variables:
            frac, whole = math.modf(float(args.get(i)) )
            if frac == 0:
                data[i] = int(args.get(i))
            else:
                data[i] = float(args.get(i)) 
            # print(i, type(i), data[i], type(data[i]))
            
        # print(data)
        my_namespace = types.SimpleNamespace()
        exec(code,  my_namespace.__dict__)
        res = my_namespace.solution(**data)
        
        context = {
            'formula': formula,
            'result': res,
        }
        return render(request, 'formula_detail.html', context)
        

class CalculateAPI(APIView):
    def post(self, request, id):
        print(request.data)
        args = request.data
        formula = Formula.objects.get(pk=id)
        code = formula.code
        data = dict()
        for i in formula.variables:
            frac, whole = math.modf(float(args.get(i)) )
            if frac == 0:
                data[i] = int(args.get(i))
            else:
                data[i] = float(args.get(i)) 
        my_namespace = types.SimpleNamespace()
        exec(code,  my_namespace.__dict__)
        res = my_namespace.solution(**data)
        return Response({'result': res})
    
    
class FunksionalHolatPage(View):
    def get(self, request):
       
        return render(request, 'funksional_holat.html')
    

class DownloadFileView(APIView):
    def post(self, request):
        if not request.user.is_authenticated:
            return Response({'message':"Yuklab olish uchun tizimga kirishingiz kerak! "}, status=status.HTTP_403_FORBIDDEN)

        input_data = request.data.get('data', [])
        full_name = request.data.get('full_name', 'unknown')
        safe_full_name = slugify(full_name) 
        data = request.data
        
        # print(request.data)
        # Split the input data into old_data and new_data
        old_data = [int(value) for i, value in enumerate(input_data) if i % 2 == 0]
        new_data = [int(value) for i, value in enumerate(input_data) if i % 2 != 0]
        # print(old_data)
        # print(new_data)
        
        excel_file_path = os.path.join(settings.BASE_DIR, 'static', 'shablon.xlsx')
        
        # Load and update the Excel file
        wb = load_workbook(excel_file_path)
        ws = wb['ФХБ']
        
        old_input = [f"D{i}" for i in range(12, 30)]
        new_input = [f"F{i}" for i in range(12, 30)]

        
        
        self.write_excel(ws, old_input, new_input, old_data, new_data)
        
        
        ws2 = wb['даража']
        old_input2 = [f"{chr(66+i)}2" for i in range(len(old_data))]
        new_input2 = [f"{chr(66+i)}3" for i in range(len(old_data))]

        self.write_excel(ws2, old_input2, new_input2, old_data, new_data)
        
        ws3 = wb['%']
        self.write_excel(ws3, old_input2, new_input2, old_data, new_data)
        
        
        
        
        ws['C2'] = data.get('full_name', '')
        ws['D3'] = data.get('birth', '')
        ws['D4'] = data.get('d_group', '')
        ws['D5'] = data.get('vtek_number', '')
        ws['C6'] = data.get('address', '')
        ws['D7'] = data.get('phone', '')
        ws['D8'] = data.get('tashxis', '')
        ws['E34'] = data.get('reabilitolog', '')
        ws['E36'] = data.get('substitute', '')
        ws['E38'] = data.get('nevrolog', '')
        ws['E40'] = data.get('vertebrolog', '')
        
        
        
        try: 
            if not os.path.exists('media/funksional_datas/'):
                os.mkdir('media/funksional_datas/')
        except:
            print("=============")
            
        time = slugify(datetime.now().strftime("%Y-%m-%d-%H-%M"))
        temp_file_name = f"funksional-xolat-{safe_full_name}-{time}.xlsx"
        file_path = os.path.join('funksional_datas',  temp_file_name)
        temp_file_path = os.path.join('media', file_path)
        wb.save(temp_file_path)

        # with open(temp_file_path, 'rb') as f:
        # print(file_path)
        funksional_data = FunksionalData(
            name=full_name,
            file=file_path  # Ensure this path is correct relative to media
        )
        funksional_data.save()

        # Create HttpResponse to return the Excel file
        with open(temp_file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={temp_file_name}'
        

        return response

    def write_excel(self, ws, old_input, new_input, old_data, new_data):
        for i in range(len(old_data)):
            if i < len(old_input):
                ws[old_input[i]] = old_data[i]
            if i < len(new_input):
                ws[new_input[i]] = new_data[i]

class InformationView(View):
    def get(self, request):
        return render(request, 'information.html')